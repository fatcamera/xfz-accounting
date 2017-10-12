"""
Label table widget.

Authors: caobinbin(caobinbin@baidu.com)
Date:    2017/10/03
"""


from __future__ import absolute_import
from __future__ import division
from __future__ import print_function
from __future__ import unicode_literals

import calendar
import datetime
import itertools
import numpy as np
import pandas as pd
import logging
import traceback
import openpyxl as px

from PyQt5 import QtCore
from PyQt5 import QtGui
from PyQt5 import QtWidgets

import datatypes
import utils
import undocommands


class DataTableModel(QtCore.QAbstractTableModel):
    """Custom table model used by label table view.
    """

    Date, Room, Source, Price, Commission, Comment = range(6)

    def __init__(self, parent=None):
        """Constructor.
        
        Keyword Arguments:
            parent {QtWidgets.QWidget} -- Widget parent. (default: {None})
        """
        super(DataTableModel, self).__init__(parent)
        self._records = pd.DataFrame([], columns=datatypes.Columns)
        self._undo_stack = QtWidgets.QUndoStack(self)

    def undo_stack(self):
        return self._undo_stack

    def set_data(self, df):
        self.beginResetModel()
        self._records = df
        self.endResetModel()

    @utils.dumpargs
    def new(self, year, month):
        days = calendar.monthrange(year, month)[1]
        date_df = pd.DataFrame(
            pd.date_range(datetime.datetime(year, month, 1), periods=days), columns=['Date'])
        room_df = pd.DataFrame(datatypes.Rooms, columns=['Room'])
        rows = itertools.product(date_df.iterrows(), room_df.iterrows())
        merge_df = pd.DataFrame(left.append(right) for (_, left), (_, right) in rows)
        merge_df = merge_df.reset_index(drop=True)

        self.set_data(
            pd.DataFrame({
                'Date': merge_df['Date'],
                'Room': merge_df['Room'],
                'Source': '',
                'Price': 0.0,
                'Commission': 0.0,
                'Comment': ''
            }, columns=datatypes.Columns)
        )
        self._undo_stack.clear()

    @utils.dumpargs
    def save(self, filename):
        success = True
        try:
            self._records.to_csv(filename, index=False, encoding='utf-8')
            self._undo_stack.setClean()
        except Exception as e:
            logging.error(traceback.format_exc())
            QtWidgets.QMessageBox.warning(
                self.parent().parent(), QtWidgets.QApplication.instance().applicationName(),
                '{}<p>{}'.format(str(e), traceback.format_exc()))
            success = False
        finally:
            return success

    @utils.dumpargs
    def open(self, filename):
        success = True
        try:
            df = pd.read_csv(filename, encoding='utf-8',
                header=0, parse_dates=[0], keep_default_na=False,
                dtype={'Source': str, 'Price': float, 'Commission': float, 'Comment': str})
            assert len(df.columns) == len(datatypes.Columns)
            assert np.all(df.columns == datatypes.Columns)
            self.set_data(df)
            self._undo_stack.clear()
        except Exception as e:
            logging.error(traceback.format_exc())
            QtWidgets.QMessageBox.warning(
                self.parent().parent(), QtWidgets.QApplication.instance().applicationName(),
                '{}<p>{}'.format(str(e), traceback.format_exc()))
            success = False
        finally:
            return success

    def _import_qunar_prepaid(self, data, filename):
        logging.info('importing {}'.format(filename))
        wb = px.load_workbook(filename)
        for row in list(wb['订单明细'].rows)[1:-1]:
            rooms = int(row[10].value)
            assert rooms == 1
            nights = int(row[11].value)
            room_nights = rooms * nights
            checkin = row[13].value
            checkout = row[14].value
            price = float(row[19].value[1:]) / room_nights
            commission = float(row[20].value[1:]) / room_nights
            room = row[25].value[:-1]
            for date in pd.date_range(checkin, checkout, closed='left'):
                data.loc[
                        (data.Date == date) & (data.Room == room),
                        ['Source', 'Price', 'Commission']
                    ] = ['去哪-预付', price, commission]

    def _import_meituan_prepaid(self, data, filename):
        logging.info('importing {}'.format(filename))
        wb = px.load_workbook(filename)
        for row in list(wb['订单详情'].rows)[1:]:
            room_nights = int(row[6].value)
            checkin, checkout = row[3].value.split('~')
            commission = float(row[8].value) / room_nights
            price = float(row[9].value) / room_nights + commission
            room = row[4].value
            for date in pd.date_range(checkin, checkout, closed='left'):
                data.loc[
                        (data.Date == date) & (data.Room == room),
                        ['Source', 'Price', 'Commission']
                    ] = ['美团-预定', price, commission]

    def _import_filename(self, data, filename):
        wb = px.load_workbook(filename, read_only=True)
        sheets = wb.get_sheet_names()
        if sheets == ['总计', '订单明细', '人工调整明细', '过期返现明细']:
            self._import_qunar_prepaid(data, filename)
        elif sheets == ['总表', '订单详情', '商家承担退款', '商家承担优惠', '调整金额']:
            self._import_meituan_prepaid(data, filename)
        else:
            raise RuntimeError('file not supported: {}'.format(filename))

    def import_files(self, filenames):
        success = True
        try:
            # import files
            temp = self._records.copy(deep=True)
            for filename in filenames:
                self._import_filename(temp, filename)
            command = undocommands.ResetCommand(self,
                QtCore.QCoreApplication.translate('DataTableModel', 'Import'))
            command.old_data = self._records.copy(deep=True)
            command.new_data = temp.copy(deep=True)
            self._undo_stack.push(command)
        except Exception as e:
            logging.error(traceback.format_exc())
            QtWidgets.QMessageBox.warning(
                self.parent().parent(), QtWidgets.QApplication.instance().applicationName(),
                '{}<p>{}'.format(str(e), traceback.format_exc()))
            success = False
        finally:
            return success

    def stat(self):
        rooms = (self._records.Source != '').sum()
        income = self._records.Price.sum()
        expense = self._records.Commission.sum()
        share = self._records[self._records.Room == '暖春'].Commission.sum()
        yanyan = self._records[self._records.Source.isin(['去哪-现付', '线下'])].Price.sum() \
            - self._records[(self._records.Source == '线下') & (self._records.Room != '暖春')].Commission.sum()
        return rooms, income, expense, share, yanyan

    def rowCount(self, index):
        """Override. Number of rows.

        Arguments:
            index {QtCore.QModelIndex} -- Model index.
        
        Returns:
            int -- Number of rows.
        """
        return self._records.shape[0]

    def columnCount(self, index):
        """Override. Number of columns.
        
        Arguments:
            index {QtCore.QModelIndex} -- Model index.
        
        Returns:
            int -- Number of columns.
        """
        return self._records.shape[1]

    def headerData(self, section, orientation, role):
        """Override. Provide header data.
        
        Arguments:
            section {int} -- Index.
            orientation {QtCore.Qt.Orientation} -- Table orientation.
            role {QtCore.Qt.ItemDataRole} -- Data role.
        
        Returns:
            QtCore.QVariant -- The data for the given role and section in the header
                                with the specified orientation.
        """
        data = None
        if role == QtCore.Qt.DisplayRole:
            if orientation == QtCore.Qt.Horizontal:
                if section == DataTableModel.Date:
                    data = QtCore.QCoreApplication.translate('DataTableModel', 'Date')
                elif section == DataTableModel.Room:
                    data = QtCore.QCoreApplication.translate('DataTableModel', 'Room')
                elif section == DataTableModel.Source:
                    data = QtCore.QCoreApplication.translate('DataTableModel', 'Source')
                elif section == DataTableModel.Price:
                    data = QtCore.QCoreApplication.translate('DataTableModel', 'Price')
                elif section == DataTableModel.Commission:
                    data = QtCore.QCoreApplication.translate('DataTableModel', 'Commission')
                elif section == DataTableModel.Comment:
                    data = QtCore.QCoreApplication.translate('DataTableModel', 'Comment')
        return data

    def data(self, index, role):
        """Override. Provide data.
        
        Arguments:
            index {QtCore.QModelIndex} -- Model index.
            role {QtCore.Qt.ItemDataRole} -- Data role.
        
        Returns:
            QtCore.QVariant -- The data stored under the given role
                                for the item referred to by the index.
        """
        data = None
        if role == QtCore.Qt.DisplayRole:
            if index.column() == DataTableModel.Date:
                data = self._records.iloc[index.row(), index.column()].strftime('%Y-%m-%d')
            elif index.column() in [DataTableModel.Price, DataTableModel.Commission]:
                data = self._records.iloc[index.row(), index.column()]
                if np.fabs(data) < 0.01:
                    data = ''
                else:
                    data = '{:.2f}'.format(data)
            else:
                data = self._records.iloc[index.row(), index.column()]
        elif role == QtCore.Qt.EditRole:
            data = self._records.iloc[index.row(), index.column()]
        elif role == QtCore.Qt.TextAlignmentRole:
            if index.column() in [DataTableModel.Date, DataTableModel.Room, DataTableModel.Source]:
                data = QtCore.Qt.AlignHCenter | QtCore.Qt.AlignVCenter
            elif index.column() in [DataTableModel.Price, DataTableModel.Commission]:
                data = QtCore.Qt.AlignRight | QtCore.Qt.AlignVCenter
            else:
                data = QtCore.Qt.AlignLeft | QtCore.Qt.AlignVCenter
        return data

    def flags(self, index):
        flag = super(DataTableModel, self).flags(index)
        if index.column() in [
                DataTableModel.Source,
                DataTableModel.Price,
                DataTableModel.Commission,
                DataTableModel.Comment
            ]:
            flag |= QtCore.Qt.ItemIsEditable
        return flag

    def setData(self, index, value, role=QtCore.Qt.EditRole):
        success = False
        #
        if index.column() == DataTableModel.Source:
            success = True
        elif index.column() in [DataTableModel.Price, DataTableModel.Commission]:
            try:
                value = float(value)
            except:
                pass
            else:
                success = True
        elif index.column() == DataTableModel.Comment:
            success = True
        #
        if success:
            command = undocommands.EditCommand(self,
                QtCore.QCoreApplication.translate('DataTableModel', 'Edit'))
            command.index = index
            command.old_value = self._records.iloc[index.row(), index.column()]
            command.new_value = value
            self._undo_stack.push(command)

        return success

    def set_value(self, index, value):
        self._records.iloc[index.row(), index.column()] = value
        self.dataChanged.emit(index, index)


class EditDelegate(QtWidgets.QStyledItemDelegate):
    """Delegate preserving background color on selection.
    """

    def __init__(self, parent=None):
        """Constructor.
        
        Keyword Arguments:
            parent {QtWidgets.QWidget} -- Widget parent. (default: {None})
        """
        super(EditDelegate, self).__init__(parent)

    def createEditor(self, parent, option, index):
        if index.column() == DataTableModel.Source:
            combo = QtWidgets.QComboBox(parent)
            combo.addItem('')
            combo.addItems(datatypes.Sources)
            return combo
        else:
            return super(EditDelegate, self).createEditor(parent, option, index)

    def setEditorData(self, editor, index):
        if index.column() == DataTableModel.Source:
            i = editor.findText(index.model().data(index, QtCore.Qt.EditRole))
            if i == -1:
                i = 0
            editor.setCurrentIndex(i)
        else:
            super(EditDelegate, self).setEditorData(editor, index)

    def setModelData(self, editor, model, index):
        if index.column() == DataTableModel.Source:
            model.setData(index, editor.currentText())
        else:
            super(EditDelegate, self).setModelData(editor, model, index)


class DataTableView(QtWidgets.QTableView):
    """Label table view.
    """

    def __init__(self, parent=None):
        """Constructor.
        
        Keyword Arguments:
            parent {QtWidgets.QWidget} -- Widget parent. (default: {None})
        """
        super(DataTableView, self).__init__(parent)
        self.setItemDelegate(EditDelegate(self))
        self.setAlternatingRowColors(True)
        self.horizontalHeader().setStretchLastSection(True)
        self.horizontalHeader().setHighlightSections(False)


class DataProxyModel(QtCore.QSortFilterProxyModel):

    def __init__(self, parent=None):
        super(DataProxyModel, self).__init__(parent)
        self._room = QtCore.QCoreApplication.translate('DataPanel', 'All')
        self._source = QtCore.QCoreApplication.translate('DataPanel', 'All')

    def filterAcceptsRow(self, source_row, source_parent):
        if self._room != QtCore.QCoreApplication.translate('DataPanel', 'All'):
            index = self.sourceModel().index(source_row, 1, source_parent)
            data = self.sourceModel().data(index, QtCore.Qt.DisplayRole)
            if data != self._room:
                return False
        if self._source != QtCore.QCoreApplication.translate('DataPanel', 'All'):
            index = self.sourceModel().index(source_row, 2, source_parent)
            data = self.sourceModel().data(index, QtCore.Qt.DisplayRole)
            if data != self._source:
                return False
        return True

    def set_room(self, room):
        self._room = room
        self.invalidateFilter()

    def set_source(self, source):
        self._source = source
        self.invalidateFilter()

    def reset_filters(self):
        self._room = QtCore.QCoreApplication.translate('DataPanel', 'All')
        self._source = QtCore.QCoreApplication.translate('DataPanel', 'All')
        self.invalidateFilter()


class DataPanel(QtWidgets.QWidget):
    """[summary]
    """

    sig_data_changed = QtCore.pyqtSignal(name='sig_data_changed')

    def __init__(self, parent=None):
        super(DataPanel, self).__init__(parent)

        filter_layout = QtWidgets.QBoxLayout(QtWidgets.QBoxLayout.LeftToRight)

        label = QtWidgets.QLabel(
            QtCore.QCoreApplication.translate('DataPanel', 'Filter Rules:'))
        label.setAlignment(QtCore.Qt.AlignLeft | QtCore.Qt.AlignVCenter)
        filter_layout.addWidget(label)

        label = QtWidgets.QLabel(
            QtCore.QCoreApplication.translate('DataPanel', 'Room'))
        label.setAlignment(QtCore.Qt.AlignRight | QtCore.Qt.AlignVCenter)
        filter_layout.addWidget(label)

        self._room_combo = QtWidgets.QComboBox()
        self._room_combo.addItem(QtCore.QCoreApplication.translate('DataPanel', 'All'))
        self._room_combo.addItems(datatypes.Rooms)
        filter_layout.addWidget(self._room_combo)

        label = QtWidgets.QLabel(
            QtCore.QCoreApplication.translate('DataPanel', 'Source'))
        label.setAlignment(QtCore.Qt.AlignRight | QtCore.Qt.AlignVCenter)
        filter_layout.addWidget(label)

        self._source_combo = QtWidgets.QComboBox()
        self._source_combo.addItem(QtCore.QCoreApplication.translate('DataPanel', 'All'))
        self._source_combo.addItems(datatypes.Sources)
        filter_layout.addWidget(self._source_combo)

        self._data_model = DataTableModel(self)
        self._proxy_model = DataProxyModel(self)
        self._proxy_model.setSourceModel(self._data_model)
        self._data_view = DataTableView()
        self._data_view.setModel(self._proxy_model)

        layout = QtWidgets.QBoxLayout(QtWidgets.QBoxLayout.TopToBottom)
        layout.addLayout(filter_layout)
        layout.addWidget(self._data_view)

        self.setLayout(layout)

        #
        self._room_combo.currentTextChanged.connect(self._proxy_model.set_room)
        self._source_combo.currentTextChanged.connect(self._proxy_model.set_source)
        self._data_model.modelReset.connect(self.sig_data_changed)
        self._data_model.dataChanged.connect(self.sig_data_changed)

    def open_bill(self, filename):
        return self._data_model.open(filename)

    def new_bill(self, year, month):
        self._data_model.new(year, month)

    def import_files(self, filenames):
        return self._data_model.import_files(filenames)

    def save_bill(self, filename):
        return self._data_model.save(filename)

    def get_stat(self):
        return self._data_model.stat()

    def undo_stack(self):
        return self._data_model.undo_stack()
